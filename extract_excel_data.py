import openpyxl
import json
import os
import re
from pprint import pprint
import shutil

class SpreadsheetParser:
    def __init__(self, filepath):
        """
        Initialize a SpreadsheetParser object.

        Parameters:
        filepath (str): The filepath of the spreadsheet.
        """
        self.workbook = openpyxl.load_workbook(filepath)

    def get_sheet_names(self):
        """
        Get the names of the sheets in the workbook.

        Returns:
        list: A list of sheet names.
        """
        return self.workbook.sheetnames

    def get_data_from_sheet(self, sheet_name):
        """
        Get the data from a sheet in the workbook.

        Parameters:
        sheet_name (str): The name of the sheet.

        Returns:
        list: A list of lists, where each inner list represents a row in the sheet and contains the values of the cells in that row.
        """
        sheet = self.workbook[sheet_name]
        data = []
        for row in sheet.rows:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            data.append(row_data)
        return data

    def get_cell_value(self, sheet_name, row, column):
        """
        Get the value of a cell in the sheet.

        Parameters:
        sheet_name (str): The name of the sheet.
        row (int): The row number of the cell (1-based).
        column (int): The column number of the cell (1-based).

        Returns:
        Any: The value of the cell.
        """
        sheet = self.workbook[sheet_name]
        return sheet.cell(row=row, column=column).value

    def get_row_values(self, sheet_name, row):
        """
        Get the values of the cells in a row in the sheet.

        Parameters:
        sheet_name (str): The name of the sheet.
        row (int): The row number (1-based).

        Returns:
        list: A list of the values of the cells in the row.
        """
        sheet = self.workbook[sheet_name]
        values = []
        for cell in sheet[row]:
            values.append(cell.value)
        return values

    def get_column_values(self, sheet_name, column):
        """
        Get the values of the cells in a column in the sheet.

        Parameters:
        sheet_name (str): The name of the sheet.
        column (int or str): The column number (1-based) or column letter.

        Returns:
        list: A list of the values of the cells in the column.
        """
        sheet = self.workbook[sheet_name]
        values = []
        for cell in sheet[column]:
            values.append(cell.value)
        return values
    
    
    def collect_data_into_dict(self, sheet_name, key_column, *value_columns):
        """
        Collect data from the sheet into a dictionary with support for multiple value columns,
        replacing any None values with 'undefined' to accommodate JavaScript usage.

        Parameters:
        sheet_name (str): The name of the sheet.
        key_column (int): The column number of the keys in the dictionary (1-based).
        value_columns (int*): A list of column numbers for the values in the dictionary (1-based).

        Returns:
        dict: A dictionary where each key (cell coordinate) maps to a list of values,
        the first being the value's label from the key column, followed by values from the specified value columns in the sheet.
        """
        sheet = self.workbook[sheet_name]
        data = {}
        for row in sheet.rows:
            key_cell = row[key_column - 1]  # Access the cell for the key
            key = f"{key_cell.coordinate}"  # Use cell coordinate as the key

            # Retrieve the label from the key cell
            label = key_cell.value if key_cell.value is not None else ''

            # Retrieve values from specified value columns, replacing None with 'undefined'
            values = [row[col - 1].value if row[col - 1].value is not None else '' for col in value_columns]
            
            # Store the label and the list of values under the key in the dictionary
            data[key] = [label] + values  # Combining the label with the list of values

        return data



    def loop_through_range(self, sheet_name, start_row, end_row, start_column, end_column):
        """
        Loop through a range of cells in the sheet.

        Parameters:
        sheet_name (str): The name of the sheet.
        start_row (int): The first row in the range (1-based).
        end_row (int): The last row in the range (1-based).
        start_column (int): The first column in the range (1-based).
        end_column (int): The last column in the range (1-based).
        """
        sheet = self.workbook[sheet_name]
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            print(row_data)

    def collect_grades_by_sheet(self, write_js=False, output_dir=""):
        """
        Collect grades data from all sheets in the workbook and write the data to JavaScript files, if requested.

        Parameters:
        write_js (bool): Whether to write the data to JavaScript files.
        output_dir (str): The output directory for the JavaScript files.
        """
        def clean_key(key):
            if not isinstance(key, str):
                key = str(key)  # Convert the key to a string if it's not already a string
            key = re.sub(r"#", "num", key)  # Replace # with num
            key = re.sub(r"[^\w\s]", "", key)  # Remove non-alphanumeric characters
            key = re.sub(r"\s+", "_", key)  # Replace spaces with underscores
            key = re.sub(r"_$", "", key)  # Remove trailing underscore, if any

            return key.lower()

        def clean_value(value):
            if not isinstance(value, str):
                value = str(value)  # Convert the value to a string if it's not already a string
            value = re.sub(r"_x000B_", "\n", value)  # Remove linebreak character

            return value

        sheet_grades = {}
        js_filenames = []  # Keep track of JS files written for each sheet

        for sheet_name in self.get_sheet_names():
            sheet = self.workbook[sheet_name]

            header_row = 1
            headers = self.get_row_values(sheet_name, header_row)
            cleaned_headers = [clean_key(header) for header in headers]

            grade_rows = self.get_data_from_sheet(sheet_name)[1:]

            sheet_grades[sheet_name] = {
                cleaned_headers[i]: [
                    {
                        "value": clean_value(row[i]),
                        "row": row_idx + header_row + 1,  # Add header_row and 1 to get the 1-based row number
                        "column": i + 1  # Convert 0-based index to 1-based column number
                    }
                    for row_idx, row in enumerate(grade_rows)
                ]
                for i in range(len(headers))
            }

            if write_js:
                if not os.path.exists(output_dir):
                    os.makedirs(output_dir)

                js_filename = os.path.join(output_dir, f"{sheet_name.replace(' ', '_')}.js")
                var_name = js_filename.split("/")[-1].split(".")[0]
                with open(js_filename, "w", encoding="utf-8") as js_file:
                    js_data = f"var {var_name} = {json.dumps(sheet_grades[sheet_name])};"
                    js_file.write(js_data)

                if sheet_name != "G_U_Wk_L" and sheet_name != "TEMPLATE_SHEET":
                    # Extract only the filename without the folder path
                    js_file_only_name = js_filename.split("/")[-1]
                    js_filenames.append(f'#include "{js_file_only_name}"')

            if write_js:
                combined_js_filename = os.path.join(output_dir, "combined.js")
                with open(combined_js_filename, "w", encoding="utf-8") as combined_js_file:
                    combined_js_file.write("\n".join(js_filenames))
                    
    
    # def write_data_to_js(self, data, output_dir, var_name='data'):
    #     """
    #     Write data to a JavaScript file formatted as a variable assignment in a specified directory
    #     and update an 'includes.js' file that references all created JavaScript files.

    #     Parameters:
    #     data (dict): The data to write to the file.
    #     output_dir (str): The directory where the .js file will be written.
    #     var_name (str): The name of the JavaScript variable.
    #     """
    #     # Ensure the output directory exists
    #     if not os.path.exists(output_dir):
    #         os.makedirs(output_dir)

    #     # Define the path for the JavaScript file
    #     file_path = os.path.join(output_dir, f"{var_name}.js")
        
    #     # Convert the Python dictionary to a JSON string
    #     js_data = json.dumps(data, indent=4)
        
    #     # Create the JavaScript code string
    #     js_code = f"var {var_name} = {js_data};"
        
    #     # Write the JavaScript code to a file
    #     with open(file_path, "w") as file:
    #         file.write(js_code)

    #     # Update the includes.js file with the new script file
    #     includes_path = os.path.join(output_dir, "includes.js")
    #     with open(includes_path, "a") as includes_file:
    #         includes_file.write(f'#include "{var_name}.js";\nvar {var_name}_DATA = {var_name};')
    
    
    def write_data_to_js(self, data, output_dir, var_name='data'):
        """
        Write data to a JavaScript file formatted as a variable assignment in a specified directory
        and update an 'includes.js' file that references all created JavaScript files.
        Ensures UTF-8 encoding and that the output directory is fresh each run.

        Parameters:
        data (dict): The data to write to the file.
        output_dir (str): The directory where the .js file will be written.
        var_name (str): The name of the JavaScript variable.
        """
        # Remove the output directory if it exists, then recreate it
        if os.path.exists(output_dir):
            shutil.rmtree(output_dir)
        os.makedirs(output_dir)

        # Define the path for the JavaScript file
        file_path = os.path.join(output_dir, f"{var_name}.js")
        
        # Convert the Python dictionary to a JSON string with UTF-8 support
        js_data = json.dumps(data, ensure_ascii=False, indent=4)
        
        # Create the JavaScript code string
        js_code = f"var {var_name} = {js_data};"
        
        # Write the JavaScript code to a file with UTF-8 encoding
        with open(file_path, "w", encoding="utf-8") as file:
            file.write(js_code)

        # Update the includes.js file with the new script file
        includes_path = os.path.join(output_dir, "includes.js")
        with open(includes_path, "w", encoding="utf-8") as includes_file:  # Using 'w' to rewrite the file each time
            includes_file.write(f'#include "{var_name}.js";\nvar {var_name}_DATA = {var_name};')


parser = SpreadsheetParser("./CA_BA_Grade_1.xlsx")
data = parser.collect_data_into_dict("G1 CA",  5, 7,8,9)
parser.write_data_to_js(data, 'output777', 'G1_CA')
